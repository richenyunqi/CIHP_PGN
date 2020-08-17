# coding=UTF-8
import openpyxl
import os


def images_count(flag, names):
    return len([i for i in names if flag in i])


def statics(terrible_path):
    final_path = terrible_path.replace('terrible', 'final')
    names = os.listdir(terrible_path)
    total_terrible_M, total_terrible_W = 0, 0
    total_M, total_W = 0, 0
    values = [[
        '日期', '需要修正（男）', '总数（男）', '需要修正所占比例（男）', '需要修正（女）', '总数（女）',
        '需要修正所占比例（女）', '需要修正（男女合计）', '不需要修正（男女合计）', '总数（男女合计）',
        '需要修正所占比例（男女合计）'
    ]]
    for n in names:
        terrible_images_path = os.path.join(terrible_path, n)
        final_dir_path = os.path.join(final_path, n)
        terrible_images_names = os.listdir(terrible_images_path)
        terrible_M = images_count('_M_', terrible_images_names) // 5
        terrible_W = images_count('_W_', terrible_images_names) // 5
        total_terrible_M += terrible_M
        total_terrible_W += terrible_W
        final_images_names = os.listdir(final_dir_path)
        final_M = images_count('_M_', final_images_names) * 90
        final_W = images_count('_W_', final_images_names) * 90
        total_M += final_M
        total_W += final_W
        values.append([
            n.replace('2020', ''), terrible_M, final_M,
            str(round(terrible_M / final_M * 100, 2)) +
            '%' if final_M != 0 else 'inf', terrible_W, final_W,
            str(round(terrible_W / final_W * 100, 2)) +
            '%' if final_W != 0 else 'inf', terrible_M + terrible_W,
            final_M + final_W - terrible_M + terrible_W, final_M + final_W,
            str(round((terrible_M + terrible_W) /
                      (final_M + final_W) * 100, 2)) + '%'
        ])
    values.append([
        'Total', total_terrible_M, total_M,
        str(round(total_terrible_M / total_M * 100, 2)) + '%',
        total_terrible_W, total_W,
        str(round(total_terrible_W / total_W * 100, 2)) + '%',
        (total_terrible_M + total_terrible_W),
        (total_M + total_W) - (total_terrible_M + total_terrible_W),
        (total_M + total_W),
        str(
            round((total_terrible_M + total_terrible_W) /
                  (total_M + total_W) * 100, 2)) + '%'
    ])
    return values


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1,
                       column=j + 1,
                       value=int(value[i][j])
                       if type(value[i][j]) == int else str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


if __name__ == '__main__':
    book_name_xlsx = 'statics.xlsx'
    sheet_name_xlsx = 'statics'
    write_excel_xlsx(book_name_xlsx, sheet_name_xlsx,
                     statics('F:/human/result/terrible'))
    read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)